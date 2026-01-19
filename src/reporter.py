import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
import os
from datetime import datetime
from .parser import CodeSnippet
from .config import ScannerConfig

class Reporter:
    def __init__(self, config: ScannerConfig, findings: List[CodeSnippet]):
        self.config = config
        self.findings = findings
        self.wb = openpyxl.Workbook()

    def bundle_code(self):
        """Extracts inline code to separate files."""
        base_dir = os.path.join(self.config.output_folder, "extracted_code")
        js_dir = os.path.join(base_dir, "inline_javascript")
        css_dir = os.path.join(base_dir, "inline_css")
        
        for d in [js_dir, css_dir]:
            if not os.path.exists(d):
                os.makedirs(d)
                
        for f in self.findings:
            if not f.full_code:
                continue
                
            # Determine extension and dir
            if f.category == 'JS':
                ext = ".js"
                target_dir = js_dir
            elif f.category == 'CSS':
                ext = ".css"
                target_dir = css_dir
            else:
                continue # External resources or others don't need bundling
                
            # Sanitize path
            rel_path = self._get_relative_path(f.file_path)
            sanitized_path = rel_path.replace(os.sep, "_").replace("/", "_").replace("\\", "_")
            
            # Sanitize type (should be clean from parser, but safety first)
            safer_type = "".join([c if c.isalnum() else "_" for c in f.code_type])
            
            # Naming convention: <dir>_<file>_<context>_line<start>-<end>.<ext>
            # Example: login_index_scriptblock_line45-78.js
            filename = f"{sanitized_path}_{safer_type}_line{f.start_line}-{f.end_line}{ext}"
            
            # Write file
            try:
                full_path = os.path.join(target_dir, filename)
                with open(full_path, "w", encoding="utf-8") as out:
                    out.write(f.full_code)
                f.bundled_file = filename
            except Exception as e:
                print(f"Failed to bundle code for {f.file_path}: {e}")

    def generate_report(self):
        # Bundle code first
        self.bundle_code()
        
        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

        # Create Tabs
        self._create_summary_sheet()
        self._create_js_sheet()
        self._create_css_sheet()
        self._create_external_sheet()

        # Save
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"InlineCode_Scan_{timestamp}.xlsx"
        output_path = os.path.join(self.config.output_folder, filename)
        
        try:
            self.wb.save(output_path)
            print(f"Report saved to: {output_path}")
        except PermissionError:
            print(f"Error: Could not save report to {output_path}. File might be open.")

    def _get_relative_path(self, absolute_path: str) -> str:
        try:
            return os.path.relpath(absolute_path, self.config.root_folder)
        except ValueError:
            return absolute_path

    def _create_summary_sheet(self):
        ws = self.wb.create_sheet("Summary")
        
        # Title
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Inline Code Detection Report"
        title_cell.font = Font(bold=True, size=16, color="2F75B5")
        
        # Metadata
        ws.cell(row=3, column=1, value="Generated:")
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        ws.cell(row=4, column=1, value="Core Path:")
        ws.cell(row=4, column=2, value=os.path.abspath(self.config.root_folder))

        # Stats
        js_count = len([f for f in self.findings if f.category == 'JS'])
        css_count = len([f for f in self.findings if f.category == 'CSS'])
        ext_count = len([f for f in self.findings if f.category == 'External'])
        total_count = len(self.findings)

        # Table Header
        ws.cell(row=6, column=1, value="Detection Summary").font = Font(bold=True, size=14)
        
        headers = ["Category", "Count"]
        header_row = 7
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="left")

        # Table Data
        data = [
            ("Inline JavaScript Findings", js_count),
            ("Inline CSS Findings", css_count),
            ("External Resources", ext_count),
            ("Total Issues", total_count)
        ]

        for i, (cat, count) in enumerate(data):
            row = header_row + 1 + i
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            # Add simple border
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = Border(bottom=Side(style='thin'))

        # Adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_js_sheet(self):
        # Columns: #, File Path, File Name, Extracted File, Type, Start Line, End Line, Code Snippet, AJAX Detected, Full Code
        headers = ["#", "File Path", "File Name", "Extracted File", "Type", "Start Line", "End Line", "Code Snippet", "AJAX Detected", "Full Code"]
        data = []
        
        js_findings = [f for f in self.findings if f.category == 'JS']
        for i, f in enumerate(js_findings, 1):
            data.append([
                i,
                f.file_path,  # Full absolute path
                os.path.basename(f.file_path), # Just filename
                f.bundled_file,
                f.code_type,
                f.start_line,
                f.end_line,
                f.snippet,
                "Yes" if f.ajax_detected else "No",
                f.full_code
            ])
        self._create_sheet("Inline JavaScript", headers, data)

    def _create_css_sheet(self):
        # Columns: #, File Path, File Name, Extracted File, Type, Start Line, End Line, Code Snippet, Full Code
        headers = ["#", "File Path", "File Name", "Extracted File", "Type", "Start Line", "End Line", "Code Snippet", "Full Code"]
        data = []
        
        css_findings = [f for f in self.findings if f.category == 'CSS']
        for i, f in enumerate(css_findings, 1):
            data.append([
                i,
                f.file_path,
                os.path.basename(f.file_path),
                f.bundled_file,
                f.code_type,
                f.start_line,
                f.end_line,
                f.snippet,
                f.full_code
            ])
        self._create_sheet("Inline CSS", headers, data)

    def _create_external_sheet(self):
        # Columns: #, File Path, File Name, Type, Resource Path, Start Line, End Line
        headers = ["#", "File Path", "File Name", "Type", "Resource Path", "Start Line", "End Line"]
        data = []
        
        ext_findings = [f for f in self.findings if f.category == 'External']
        for i, f in enumerate(ext_findings, 1):
            data.append([
                i,
                f.file_path,
                os.path.basename(f.file_path),
                f.code_type,
                f.snippet, # snippet holds the URL for external
                f.start_line,
                f.end_line
            ])
        self._create_sheet("External Resources", headers, data)

    def _create_sheet(self, title: str, headers: List[str], data_rows: List[List[str]]):
        ws = self.wb.create_sheet(title)
        
        # Header Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Write Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # Write Data
        for row_num, row_data in enumerate(data_rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                val_str = str(cell_value)
                # Excel cell limit is 32767 chars
                if len(val_str) > 32000:
                    val_str = val_str[:32000] + "..."
                cell.value = val_str
                
                # Align top for readability
                cell.alignment = Alignment(vertical="top", wrap_text=True if len(val_str) > 50 else False)
                cell.border = thin_border

        # Adjust column widths (basic heuristic)
        for col_num, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = col[0].column_letter
            
            # Don't auto-expand "Full Code" or "Snippet" too much
            header_val = ws.cell(row=1, column=col_num).value
            if header_val in ["Code Snippet", "Full Code", "Resource Path"]:
                ws.column_dimensions[column_letter].width = 50
                continue
                
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            if adjusted_width < 10: adjusted_width = 10
            ws.column_dimensions[column_letter].width = adjusted_width
