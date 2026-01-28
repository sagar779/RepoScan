import os
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================================
# MODULE 1: LOGIC & ASSESSMENT - Ternary Classification System
# ============================================================================

class DirectoryStructureError(Exception):
    """Raised when the extracted_code directory doesn't match expected schema."""
    pass


def validate_directory_structure(extracted_path):
    """
    Validates that the extraction directory follows the expected schema:
    - extracted_code/js/internal/
    - extracted_code/js/inline/
    - extracted_code/css/internal/
    - extracted_code/css/inline/
    """
    required_dirs = [
        os.path.join(extracted_path, 'js', 'internal'),
        os.path.join(extracted_path, 'js', 'inline'),
        os.path.join(extracted_path, 'css', 'internal'),
        os.path.join(extracted_path, 'css', 'inline')
    ]
    
    missing = [d for d in required_dirs if not os.path.exists(d)]
    
    if missing:
        raise DirectoryStructureError(
            f"Invalid directory structure. Missing required folders:\n" +
            "\n".join(f"  - {os.path.relpath(d, extracted_path)}" for d in missing) +
            f"\n\nExpected structure:\n"
            f"  {extracted_path}/\n"
            f"    js/internal/  (Script blocks)\n"
            f"    js/inline/    (Inline event handlers)\n"
            f"    css/internal/ (Style blocks)\n"
            f"    css/inline/   (Inline style attributes)"
        )


def parse_metadata(filename):
    """
    Parses original file info from the extracted filename.
    Format: OriginalPath_Type_lineStart-End.ext
    Returns: (OriginalPath, LineRange, Type)
    """
    try:
        # Match the standard format defined in the tool
        # Example: Views_Home_Index.cshtml_scriptblock_line10-25.js
        match = re.search(r'(.+)_([a-zA-Z0-9]+)_line(\d+)-(\d+)\.(js|css)$', filename)
        if match:
            sanitized, code_type, start, end, ext = match.groups()
            orig_path = sanitized.replace('_', '/')
            return orig_path, f"{start}-{end}", code_type, int(start), int(end)
    except:
        pass
    return filename, "Unknown", "Unknown", 0, 0


def detect_razor_patterns(content):
    """
    Detects Razor syntax patterns in code.
    Returns: (state, razor_tokens, pattern_type)
    
    State A (Green): No Razor syntax
    State B (Yellow): Razor syntax in strings/values (bridgeable)
    State C (Red): Razor control flow (blocked)
    """
    # Control Flow Patterns (State C - Blocked)
    control_flow_patterns = [
        r'@if\s*\(',
        r'@else',
        r'@foreach\s*\(',
        r'@for\s*\(',
        r'@while\s*\(',
        r'@switch\s*\(',
        r'@using\s*\(',
        r'<%\s*if',
        r'<%\s*for',
        r'<%\s*while'
    ]
    
    for pattern in control_flow_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            return 'C', [], 'control_flow'
    
    # Value/String Patterns (State B - Bridgeable)
    value_patterns = [
        r'@Url\.Action\([^)]+\)',
        r'@Url\.Content\([^)]+\)',
        r'@Model\.\w+',
        r'@ViewBag\.\w+',
        r'@ViewData\[[^\]]+\]',
        r'@Html\.\w+',
        r'<%=\s*[^%]+%>'
    ]
    
    razor_tokens = []
    for pattern in value_patterns:
        matches = re.findall(pattern, content)
        if matches:
            razor_tokens.extend(matches)
    
    if razor_tokens:
        return 'B', razor_tokens, 'value_injection'
    
    # No Razor syntax found (State A - Ready)
    return 'A', [], 'clean'


def detect_ajax_calls(content):
    """
    Detects AJAX patterns in JavaScript code.
    Returns: True if AJAX call is found
    """
    ajax_patterns = [
        r'\$\.ajax\s*\(',
        r'\$\.get\s*\(',
        r'\$\.post\s*\(',
        r'\.ajax\s*\(',
        r'fetch\s*\(',
        r'XMLHttpRequest\s*\(',
        r'\$http\.',
        r'axios\.'
    ]
    
    for pattern in ajax_patterns:
        if re.search(pattern, content, re.IGNORECASE):
            return True
    return False


def classify_code(filename, content, file_category):
    """
    Classifies code into State A/B/C with AJAX detection.
    
    Args:
        filename: Name of the extracted file
        content: Code content
        file_category: 'internal' or 'inline'
    
    Returns: (state, complexity, status, action, reason, razor_count, is_ajax)
    """
    # Parse metadata
    orig_path, lines, code_type, start_line, end_line = parse_metadata(filename)
    
    # Detect Razor patterns
    state, razor_tokens, pattern_type = detect_razor_patterns(content)
    
    # Detect AJAX
    is_ajax = detect_ajax_calls(content)
    
    # CSS-specific handling
    if filename.endswith('.css'):
        # CSS rarely has complex Razor logic, but can have URL references
        if state == 'B':
            # Simple image paths are OK, complex logic is manual
            has_complex_razor = any('@' in token and '(' in token for token in razor_tokens)
            if has_complex_razor:
                state = 'C'
                return (state, "Low", "Blocked", "Manual Intervention Required", 
                        "CSS contains complex Razor logic", len(razor_tokens), False)
            else:
                return (state, "Medium", "Bridgeable", "Auto-Refactor via Bridge",
                        "CSS contains simple Razor URL references", len(razor_tokens), False)
        elif state == 'C':
            return (state, "Low", "Blocked", "Manual Intervention Required",
                    "CSS contains Razor control flow", len(razor_tokens), False)
        else:
            return (state, "High", "Ready", "Externalized - Clean CSS",
                    "No server-side logic detected", 0, False)
    
    # JavaScript classification
    if state == 'A':
        # Green - Ready
        return (state, "High", "Ready", "Externalized - Clean Code",
                "No server-side logic detected", 0, is_ajax)
    
    elif state == 'B':
        # Yellow - Bridgeable
        if is_ajax:
            return (state, "Medium", "Bridgeable", "Auto-Refactor via Bridge",
                    f"AJAX call with {len(razor_tokens)} Razor tokens - Bridge Pattern applicable",
                    len(razor_tokens), is_ajax)
        else:
            return (state, "Medium", "Bridgeable", "Auto-Refactor via Bridge",
                    f"Contains {len(razor_tokens)} Razor value tokens - Bridge Pattern applicable",
                    len(razor_tokens), is_ajax)
    
    else:  # state == 'C'
        # Red - Blocked
        return (state, "Low", "Blocked", "Manual Intervention Required",
                "Contains Razor control flow - cannot auto-refactor",
                len(razor_tokens), is_ajax)


def analyze_folder(folder_path, sheet, file_type, category, metrics):
    """
    Scans a folder and populates the Excel sheet with enhanced metadata.
    
    Args:
        folder_path: Path to the folder to analyze
        sheet: Excel worksheet object
        file_type: 'JS' or 'CSS'
        category: 'internal' or 'inline'
        metrics: Dictionary to track refactorability metrics
    """
    # Enhanced Header
    headers = ["Original File", "Lines", "Code Type", "Category", "State", 
               "Complexity", "Status", "Action", "Reason", "Razor Tokens", 
               "AJAX", "Snippet"]
    sheet.append(headers)
    
    # Style Header
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    widths = [35, 12, 15, 12, 8, 12, 15, 30, 45, 12, 8, 50]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64+i)].width = width
    
    if not os.path.exists(folder_path):
        return
    
    for root, _, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            except Exception as e:
                print(f"Error reading {file}: {e}")
                continue
            
            # Parse Metadata
            orig_path, lines, code_type, start_line, end_line = parse_metadata(file)
            
            # Classify
            state, complexity, status, action, reason, razor_count, is_ajax = classify_code(file, content, category)
            
            # Update metrics
            metrics['total'] += 1
            if state == 'A':
                metrics['clean'] += 1
            elif state == 'B':
                metrics['bridge'] += 1
            if category == 'inline' and state in ['A', 'B']:
                metrics['inline'] += 1
            
            # Create snippet
            snippet = content[:150].replace('\n', ' ').strip()
            
            # Build row
            row = [orig_path, lines, code_type, category, state, complexity, 
                   status, action, reason, razor_count, 
                   "Yes" if is_ajax else "No", snippet]
            sheet.append(row)
            
            # Color coding based on State
            row_idx = sheet.max_row
            state_cell = sheet.cell(row=row_idx, column=5)
            status_cell = sheet.cell(row=row_idx, column=7)
            
            if state == 'C':  # Blocked - Red
                fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                font = Font(color="9C0006", bold=True)
            elif state == 'B':  # Bridgeable - Yellow
                fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                font = Font(color="9C6500", bold=True)
            else:  # Ready - Green
                fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                font = Font(color="006100", bold=True)
            
            state_cell.fill = fill
            state_cell.font = font
            status_cell.fill = fill


def create_summary_sheet(wb, metrics):
    """
    Creates a summary sheet with KPI and Refactorability Index.
    """
    ws = wb.create_sheet("Summary & KPIs", 0)
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "Refactoring Assessment - Summary Report"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Metrics
    ws.append([])
    ws.append(["Metric", "Count", "Weight", "Description"])
    
    # Style metrics header
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    data = [
        ["Total Extracted Instances", metrics['total'], "1.0", "All extracted code blocks"],
        ["State A (Clean/Ready)", metrics['clean'], "1.0", "No Razor syntax - safe to externalize"],
        ["State B (Bridgeable)", metrics['bridge'], "1.0", "Razor values - Bridge Pattern applicable"],
        ["Inline Handlers Converted", metrics['inline'], "0.8", "Event handlers moved to addEventListener"],
    ]
    
    for row in data:
        ws.append(row)
    
    # Calculate Refactorability Index
    total = metrics['total']
    if total > 0:
        refactor_percentage = ((metrics['clean'] + metrics['bridge'] + (metrics['inline'] * 0.8)) / total) * 100
    else:
        refactor_percentage = 0
    
    # Add formula row
    ws.append([])
    ws.append(["Refactorability Index", f"{refactor_percentage:.2f}%", "", 
               "Weighted success rate of auto-refactoring"])
    
    # Style the index row
    index_row = ws.max_row
    ws.cell(index_row, 1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(index_row, 1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(index_row, 2).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(index_row, 2).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(index_row, 4).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    
    return refactor_percentage


def generate_report(extracted_path, output_path):
    """
    Generates the enhanced Refactoring Assessment Excel report.
    """
    # Validate directory structure first
    validate_directory_structure(extracted_path)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Initialize metrics tracker
    metrics = {'total': 0, 'clean': 0, 'bridge': 0, 'inline': 0}
    
    # JavaScript Sheet
    ws_js = wb.active
    ws_js.title = "JavaScript Assessment"
    js_path = os.path.join(extracted_path, "inline_javascript")
    analyze_folder(js_path, ws_js, "JS")
    
    # CSS Inline (Attributes)
    ws_css_inl = wb.create_sheet("CSS - Inline Styles")
    css_inl_path = os.path.join(extracted_path, "css", "inline")
    analyze_folder(css_inl_path, ws_css_inl, "CSS", "inline", metrics)
    
    # Create Summary Sheet
    refactor_percentage = create_summary_sheet(wb, metrics)
    
    # Save
    try:
        wb.save(output_path)
        print(f"\n{'='*70}")
        print(f"[OK] Assessment Report Generated: {output_path}")
        print(f"{'='*70}")
        print(f"Refactorability Index: {refactor_percentage:.2f}%")
        print(f"  * Total Instances: {metrics['total']}")
        print(f"  * Clean (State A): {metrics['clean']}")
        print(f"  * Bridgeable (State B): {metrics['bridge']}")
        print(f"  * Inline Handlers: {metrics['inline']}")
        print(f"{'='*70}\n")
    except Exception as e:
        print(f"[X] Failed to save report: {e}")


def main():
    parser = argparse.ArgumentParser(
        description="Analyze extracted code for refactorability with State-Logic Decoupling."
    )
    parser.add_argument("--extracted", required=True, 
                        help="Path to 'extracted_code' folder")
    parser.add_argument("--output", required=False, 
                        help="Path for the output Excel file (default: parent of extracted folder)")
    
    args = parser.parse_args()
    
    # Default output path if not specified
    if not args.output:
        parent_dir = os.path.dirname(os.path.normpath(args.extracted))
        args.output = os.path.join(parent_dir, "Refactoring_Assessment.xlsx")
    
    try:
        generate_report(args.extracted, args.output)
    except DirectoryStructureError as e:
        print(f"\n[X] Directory Structure Error:\n{e}\n")
        exit(1)
    except Exception as e:
        print(f"\n[X] Unexpected Error: {e}\n")
        exit(1)


if __name__ == "__main__":
    main()
