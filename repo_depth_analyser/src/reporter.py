import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class Reporter:
    def __init__(self, output_dir):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def clean_old_reports(self):
        """Removes existing tracker files to ensure a clean output."""
        pattern = os.path.join(self.output_dir, "Application_Depth_Tracker_*.xlsx")
        for f in glob.glob(pattern):
            try:
                os.remove(f)
                print(f"Removed old report: {f}")
            except Exception as e:
                print(f"Error removing {f}: {e}")

    def _style_worksheet(self, ws):
        """Applies professional styling to a worksheet."""
        # Header Style: Bold, Blue Background, White Text
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply to header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Auto-adjust column widths (optimized for performance)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            column_name = col[0].value
            
            # Sample only first 100 rows for width calculation (performance optimization)
            sample_size = min(100, len(col))
            for cell in list(col)[:sample_size]:
                try:
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.05
            
            # Special handling for Full_Path column - allow wider width
            if column_name == 'Full_Path':
                ws.column_dimensions[column].width = min(adjusted_width, 100)
            else:
                ws.column_dimensions[column].width = min(adjusted_width, 50)

        # Add AutoFilter
        ws.auto_filter.ref = ws.dimensions

    def generate_report(self, inventory, dir_stats, ajax_details=None):
        """Generates a standardized Excel report."""
        self.clean_old_reports()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(self.output_dir, f"Application_Depth_Tracker_{timestamp}.xlsx")
        
        # --- Prepare Data ---

        # 1. Summary Dashboard Data
        df_inv = pd.DataFrame(inventory)
        total_files = len(df_inv)
        total_lines = df_inv['Line_Count'].sum() if not df_inv.empty else 0
        total_size_mb = df_inv['Size_KB'].sum() / 1024 if not df_inv.empty else 0

        # Extension Breakdown
        if not df_inv.empty:
            ext_counts = df_inv['Extension'].value_counts().reset_index()
            ext_counts.columns = ['Extension', 'File Count']
        else:
            ext_counts = pd.DataFrame(columns=['Extension', 'File Count'])

        # Calculate Complexity Metrics Totals
        if not df_inv.empty:
            total_inline_css = df_inv['Inline_CSS_Count'].sum()
            total_internal_style = df_inv['Internal_Style_Blocks_Count'].sum()
            total_external_css = df_inv['External_Stylesheet_Links_Count'].sum()
            total_inline_js = df_inv['Inline_JS_Count'].sum()
            total_internal_script = df_inv['Internal_Script_Blocks_Count'].sum()
            total_external_js = df_inv['External_Script_Tags_Count'].sum()
            total_ajax = df_inv['AJAX_Calls_Count'].sum()
            files_with_ajax = len(df_inv[df_inv['Has_Ajax_Calls'] == 'Yes'])
            total_dynamic_js = df_inv['Dynamic_JS_Gen_Count'].sum()
            total_dynamic_css = df_inv['Dynamic_CSS_Gen_Count'].sum()
        else:
            total_inline_css = total_internal_style = total_external_css = 0
            total_inline_js = total_internal_script = total_external_js = 0
            total_ajax = files_with_ajax = total_dynamic_js = total_dynamic_css = 0

        # Create Summary DataFrame with Basic Metrics
        summary_data = {
            'Metric': ['Total Files', 'Total Code Lines', 'Total Size (MB)'],
            'Value': [total_files, total_lines, round(total_size_mb, 2)]
        }
        df_summary_main = pd.DataFrame(summary_data)
        
        # Create Complexity Metrics Summary
        complexity_summary_data = {
            'Category': [
                '--- CSS Patterns ---',
                'Inline CSS (style="...")',
                'Internal Style Blocks (<style>)',
                'External Stylesheets (<link>)',
                '',
                '--- JavaScript Patterns ---',
                'Inline JS (event handlers)',
                'Internal Script Blocks (<script>)',
                'External Script Tags (src="...")',
                '',
                '--- AJAX & Network Calls ---',
                'Total AJAX Calls Detected',
                'Files with AJAX',
                '',
                '--- Dynamic Code Generation ---',
                'Dynamic JS (eval, innerHTML, etc.)',
                'Dynamic CSS (style manipulation)'
            ],
            'Count': [
                '',
                total_inline_css,
                total_internal_style,
                total_external_css,
                '',
                '',
                total_inline_js,
                total_internal_script,
                total_external_js,
                '',
                '',
                total_ajax,
                files_with_ajax,
                '',
                '',
                total_dynamic_js,
                total_dynamic_css
            ]
        }
        df_complexity_summary = pd.DataFrame(complexity_summary_data)

        # 2. Directory Analysis Data
        dir_rows = []
        for dirname, stats in dir_stats.items():
            # Format extension breakdown string
            ext_str = ", ".join([f"{k}: {v}" for k, v in stats.get('extensions', {}).items()])
            
            dir_rows.append({
                'Directory': dirname,
                'Depth': stats.get('depth', 0),
                'File_Count': stats['count'],
                'Total_Lines': stats['lines'],
                'Extensions_Breakdown': ext_str
            })
        df_dir_stats = pd.DataFrame(dir_rows)
        # Sort by depth then name
        if not df_dir_stats.empty:
            df_dir_stats = df_dir_stats.sort_values(by=['Depth', 'Directory'])

        # 3. File Inventory Data (Basic Metadata Only)
        cols_basic = ['Directory', 'Filename', 'Extension', 'Line_Count', 'Size_KB', 'Full_Path']
        df_details = df_inv[cols_basic] if not df_inv.empty else pd.DataFrame(columns=cols_basic)
        
        # 4. Complexity Metrics Data (Dedicated Tab)
        cols_complexity = [
            'Full_Path', 'Filename',
            'Inline_CSS_Count', 'Internal_Style_Blocks_Count', 'External_Stylesheet_Links_Count',
            'Inline_JS_Count', 'Internal_Script_Blocks_Count', 'External_Script_Tags_Count', 
            'AJAX_Calls_Count', 'Has_Ajax_Calls', 'Dynamic_JS_Gen_Count', 'Dynamic_CSS_Gen_Count'
        ]
        df_complexity = df_inv[cols_complexity] if not df_inv.empty else pd.DataFrame(columns=cols_complexity)

        # 5. AJAX Detailed Report Data
        df_ajax = pd.DataFrame(ajax_details) if ajax_details else pd.DataFrame(columns=['File_Path', 'Line', 'Code_Snippet', 'Category', 'Capability', 'Difficulty'])

        # Rename Directory to Folder Path for File_Details tab
        df_details = df_details.rename(columns={'Directory': 'Folder Path'})

        # --- Write to Excel ---
        print(f"Generating report at {output_file}...")
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Tab 1: Summary_Dashboard
                # Write Summary Metrics starting from row 0
                df_summary_main.to_excel(writer, sheet_name='Summary_Dashboard', startrow=0, startcol=0, index=False)
                
                # Write Complexity Metrics Summary
                start_row_complexity = len(df_summary_main) + 3
                ws = writer.book['Summary_Dashboard']
                ws.cell(row=start_row_complexity, column=1, value="Complexity Metrics Summary").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=start_row_complexity, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                df_complexity_summary.to_excel(writer, sheet_name='Summary_Dashboard', startrow=start_row_complexity, startcol=0, index=False)
                
                # Write Extension Breakdown below
                start_row_ext = start_row_complexity + len(df_complexity_summary) + 3
                ws.cell(row=start_row_ext, column=1, value="Global Extension Breakdown").font = Font(bold=True, size=12, color="FFFFFF")
                ws.cell(row=start_row_ext, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                ext_counts.to_excel(writer, sheet_name='Summary_Dashboard', startrow=start_row_ext, startcol=0, index=False)
                
                # Tab 2: Directory_Analysis
                df_dir_stats.to_excel(writer, sheet_name='Directory_Analysis', index=False)

                # Tab 3: File_Details
                df_details.to_excel(writer, sheet_name='File_Details', index=False)
                
                # Tab 4: Complexity_Metrics
                df_complexity.to_excel(writer, sheet_name='Complexity_Metrics', index=False)
                
                # Tab 5: AJAX_Detailed_Report
                df_ajax.to_excel(writer, sheet_name='AJAX_Detailed_Report', index=False)

                # Apply Styling
                self._style_worksheet(writer.sheets['Summary_Dashboard'])
                self._style_worksheet(writer.sheets['Directory_Analysis'])
                self._style_worksheet(writer.sheets['File_Details'])
                self._style_worksheet(writer.sheets['Complexity_Metrics'])
                
                self._style_worksheet(writer.sheets['AJAX_Detailed_Report'])

            print("Report generated successfully.")
            return output_file
        except Exception as e:
            print(f"Failed to generate report: {e}")
            return None
