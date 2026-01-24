import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List
import os
import logging
from datetime import datetime
from .parser import CodeSnippet
from .config import ScannerConfig

class Reporter:
    def __init__(self, config: ScannerConfig, findings: List[CodeSnippet]):
        self.config = config
        self.findings = findings
        self.wb = openpyxl.Workbook()

    def bundle_code(self):
        """Extracts inline code to separate files with granular organization."""
        base_dir = os.path.join(self.config.output_folder, "extracted_code")
        
        # Granular Folders
        folders = {
            "inline_js": os.path.join(base_dir, "inline_js"),
            "internal_js": os.path.join(base_dir, "internal_js"),
            "inline_css": os.path.join(base_dir, "inline_css"),
            "internal_css": os.path.join(base_dir, "internal_css")
        }
        
        for d in folders.values():
            if not os.path.exists(d):
                os.makedirs(d)
                
        for f in self.findings:
            if not f.full_code:
                continue
                
            # Determine target folder and extension
            target_folder = ""
            ext = ""
            
            if f.category == 'JS':
                ext = ".js"
                if f.source_type == 'LOCAL': target_folder = folders["internal_js"]
                else: target_folder = folders["inline_js"] # Inline & Remote-but-inline-context
            elif f.category == 'CSS':
                ext = ".css"
                if f.source_type == 'LOCAL': target_folder = folders["internal_css"]
                else: target_folder = folders["inline_css"]
            else:
                continue 

            # Filename Convention: Full Path Structure to avoid collisions
            # Format: {Path_Structure}_{Type}_L{Line}.ext
            # e.g. Views_Home_Index_cshtml_scriptblock_L45.js
            
            rel_path = self._get_relative_path(f.file_path)
            # Sanitize path: Replace separators and dots (except strict extension if needed)
            safe_path = rel_path.replace(":", "").replace(os.sep, "_").replace("/", "_").replace("\\", "_").replace(".", "_")
            
            safe_type = "".join([c if c.isalnum() else "_" for c in f.code_type])
            filename = f"{safe_path}_{safe_type}_L{f.start_line}{ext}"
            
            # Write file
            try:
                full_path = os.path.join(target_folder, filename)
                with open(full_path, "w", encoding="utf-8") as out:
                    out.write(f.full_code)
                f.bundled_file = filename
            except Exception as e:
                logging.error(f"Failed to bundle code for {f.file_path}: {e}")

    def generate_report(self):
        # Bundle code first
        self.bundle_code()
        
        # 1. Code Inventory Tracker (Now includes AJAX)
        self._create_inventory_tracker()
        
        # 2. Refactoring & Extraction Tracker (Split Tabs)
        self._create_refactoring_tracker()
        
        # 3. Crawler Input Tracker
        self._create_crawler_tracker()

    def _create_inventory_tracker(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.remove(ws)
        
        self.wb = wb 
        self._create_summary_sheet()
        
        # JS Trackers (Refactored)
        self._create_inline_js_sheet()
        self._create_internal_js_sheet()
        self._create_external_js_sheet()
        
        # CSS Trackers (Refactored)
        self._create_inline_css_sheet()
        self._create_internal_css_sheet()
        self._create_external_css_sheet()
        
        self._create_ajax_sheet() # Merged back
        self._create_legend_sheet() # New Legend
        
        self._save_wb(wb, "Code_Inventory.xlsx")

    # Removed _create_ajax_tracker as it is merged

    def _create_refactoring_tracker(self):
        wb = openpyxl.Workbook()
        self.wb = wb
        wb.remove(wb.active)
        # Split into JS and CSS
        self._create_refactoring_sheet("JS")
        self._create_refactoring_sheet("CSS")
        self._save_wb(wb, "Refactoring_Tracker.xlsx")

    def _create_crawler_tracker(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Crawler Input")
        wb.remove(wb.active if "Sheet" in wb.sheetnames else wb.worksheets[0])
        
        headers = ["Target URL", "Source File", "Rationale", "Interaction Hints"]
        data = []
        
        # Filter for HTML/ASPX files
        seen_files = set()
        for f in self.findings:
            if f.file_path in seen_files: continue
            
            ext = os.path.splitext(f.file_path)[1].lower()
            if ext in ['.html', '.htm', '.aspx', '.cshtml', '.php', '.jsp']:
                # Transform path to localhost URL (Assumption/Placeholder)
                rel_path = self._get_relative_path(f.file_path).replace("\\", "/")
                target = f"http://localhost/{rel_path}"
                
                rationale = "Page Entry Point"
                hints = "Check CSP"
                if any(x.ajax_detected for x in self.findings if x.file_path == f.file_path):
                    rationale += ", Contains AJAX"
                
                data.append([target, f.file_path, rationale, hints])
                seen_files.add(f.file_path)
                
        # Basic Write (inline here for simplicity as it's new)
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)
            
        for row, row_data in enumerate(data, 2):
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=val)
                
        self._save_wb(wb, "Crawler_Input.xlsx")

    def _save_wb(self, wb, filename):
        output_path = os.path.join(self.config.output_folder, filename)
        try:
            wb.save(output_path)
            logging.info(f"Report saved to: {output_path}")
        except PermissionError:
            logging.error(f"Could not save report to {output_path}. File might be open.")

    def _get_relative_path(self, absolute_path: str) -> str:
        try:
            return os.path.relpath(absolute_path, self.config.root_folder)
        except ValueError:
            return absolute_path

    def _create_summary_sheet(self):
        ws = self.wb.create_sheet("Summary")
        
        # Title
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Inline Code Detection Report"
        title_cell.font = Font(bold=True, size=16, color="2F75B5")
        
        # Metadata
        ws.cell(row=3, column=1, value="Generated:")
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        ws.cell(row=4, column=1, value="Core Path:")
        ws.cell(row=4, column=2, value=os.path.abspath(self.config.root_folder))

        # Stats Breakdown
        
        # 1. Inline JS (Attributes: onclick, javascript:) -> Matches User "Inline"
        inline_js_attr = len([f for f in self.findings if f.category == 'JS' and f.source_type == 'INLINE' and f.code_type != 'scriptblock'])
        
        # 2. Internal JS (Script Blocks) -> Matches User "Internal"
        internal_js_blocks = len([f for f in self.findings if f.category == 'JS' and f.source_type == 'INLINE' and f.code_type == 'scriptblock'])
        
        # 3. External JS (All src references: Local files AND Remote URLs) -> Matches User "External"
        # Parser labels local src as 'source_type=LOCAL'. Remote as 'REMOTE'.
        external_js_refs = len([f for f in self.findings if f.category in ['JS', 'External', 'Internal'] and f.source_type in ['LOCAL', 'REMOTE']])
        # Note: We filter strictly just in case category names vary, but source_type is the source of truth here.
        # Actually, let's be precise based on 'JS'/'External' categories in parser:
        # local src -> category='Internal', source='LOCAL'
        # remote src -> category='External', source='REMOTE'
        # We want to group BOTH as "External" in the report.
        external_js_combined = len([f for f in self.findings if f.source_type in ['LOCAL', 'REMOTE'] and ('script' in f.code_type.lower() or f.category in ['JS', 'Internal', 'External']) and 'css' not in f.code_type.lower() and 'style' not in f.code_type.lower()])

        # CSS Logic
        inline_css_attr = len([f for f in self.findings if f.category == 'CSS' and f.source_type == 'INLINE' and 'styleblock' not in f.code_type])
        internal_css_blocks = len([f for f in self.findings if f.category == 'CSS' and f.source_type == 'INLINE' and f.code_type == 'styleblock'])
        
        # External CSS (Local <link> + Remote <link>)
        external_css_combined = len([f for f in self.findings if f.source_type in ['LOCAL', 'REMOTE'] and ('style' in f.code_type.lower() or 'css' in f.code_type.lower() or f.category == 'CSS')])

        # Explicit sum
        total_count = inline_js_attr + internal_js_blocks + external_js_combined + inline_css_attr + internal_css_blocks + external_css_combined
        
        # Debug
        if len(self.findings) != total_count:
            # It's possible some finding falls through if code_type/category is weird.
            logging.debug(f"Note: Total findings ({len(self.findings)}) != Displayed Sum ({total_count}).")
        
        # AJAX & Dynamic Stats (Literal Counts)
        ajax_count = sum([getattr(f, 'ajax_count', 0) for f in self.findings])
        inline_ajax = sum([getattr(f, 'ajax_count', 0) for f in self.findings if f.ajax_detected and f.is_inline_ajax])
        external_ajax = sum([getattr(f, 'ajax_count', 0) for f in self.findings if f.ajax_detected and not f.is_inline_ajax])
        server_deps = sum([getattr(f, 'ajax_count', 0) for f in self.findings if f.ajax_detected and f.has_server_deps])
        clean_ajax = sum([getattr(f, 'ajax_count', 0) for f in self.findings if f.ajax_detected and f.is_inline_ajax and not f.has_server_deps])
        
        dynamic_count = sum([getattr(f, 'dynamic_count', 0) for f in self.findings])

        # Table Header
        ws.cell(row=6, column=1, value="Detection Summary").font = Font(bold=True, size=14)
        
        headers = ["Category", "Count", "Criteria / Reference"]
        header_row = 7
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="left")

        # Table Data
        data = [
            ("Inline JS (Attributes)", inline_js_attr, "Direct attributes: onclick, onload, href='javascript:...'"),
            ("Internal JS (Blocks)", internal_js_blocks, "Embedded blocks: <script>...</script>"),
            ("External JS (Files)", external_js_combined, "References: <script src='...'> (Local & Remote)"),
            ("", "", ""),
            ("Inline CSS (Attributes)", inline_css_attr, "Direct attributes: style='...'"),
            ("Internal CSS (Blocks)", internal_css_blocks, "Embedded blocks: <style>...</style>"),
            ("External CSS (Files)", external_css_combined, "References: <link rel='stylesheet'> (Local & Remote)"),
            ("", "", ""),
            ("Total Issues", total_count, "Sum of all findings"),
            ("", "", ""),  # Spacer
            ("Total AJAX Calls Found", ajax_count, "Regex match for $.ajax, fetch, xhr, axios, etc."),
            ("  - Inline AJAX", inline_ajax, "AJAX patterns found inside <script> blocks"),
            ("  - External AJAX", external_ajax, "AJAX patterns found inside local .js files"),
            ("  - With Server Dependencies", server_deps, "AJAX containing @Model, @ViewBag, or ASP tags"),
            ("  - Clean/Extractable", clean_ajax, "AJAX with no server-side dependency markers"),
            ("", "", ""),
            ("Total Dynamic Code Sinks Found", dynamic_count, "Regex match for eval(), innerHTML, document.write()"),
        ]

        for i, (cat, count, criteria) in enumerate(data):
            row = header_row + 1 + i
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=criteria)
            # Add simple border
            for col in [1, 2, 3]:
                ws.cell(row=row, column=col).border = Border(bottom=Side(style='thin'))

        # Adjust widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['C'].width = 60

    # --- JS Sheets ---
    def _create_inline_js_sheet(self):
        """1. Inline JS: Attributes (onclick, etc.)"""
        headers = ["File Path", "File Name", "Context", "Line", "Code Snippet", "Full Code"]
        data = []
        # Filter: JS + Inline Source + NOT Script Block
        findings = [f for f in self.findings if f.category == 'JS' and f.source_type == 'INLINE' and f.code_type != 'scriptblock']
        
        for f in findings:
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.code_type,
                f.start_line,
                f.snippet,
                f.full_code
            ])
        self._create_sheet("Inline JS (Attributes)", headers, data)

    def _create_internal_js_sheet(self):
        """2. Internal JS: Script Blocks"""
        headers = ["File Path", "File Name", "Extracted File", "Line", "Length (Lines)", "AJAX?", "Code Snippet", "Full Code"]
        data = []
        # Filter: JS + Inline Source + IS Script Block
        findings = [f for f in self.findings if f.category == 'JS' and f.source_type == 'INLINE' and f.code_type == 'scriptblock']
        
        for f in findings:
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.bundled_file,
                f.start_line,
                (f.end_line - f.start_line),
                "Yes" if f.ajax_detected else "No",
                f.snippet,
                f.full_code
            ])
        self._create_sheet("Internal JS (Blocks)", headers, data)

    def _create_external_js_sheet(self):
        """3. External JS: Local Files & Remote References"""
        headers = ["File Path", "File Name", "Reference Type", "Source/URL", "Line", "Is Remote?"]
        data = []
        
        # Filter: 
        # a) Category=JS + Source=LOCAL (Standalone .js file)
        # b) Category=Internal (Local <script src>)
        # c) Category=External (Remote <script src>)
        # AND 'script' keyword check for Internal/External to filter out CSS
        
        findings = [f for f in self.findings if 
                    (f.category == 'JS' and f.source_type == 'LOCAL') or 
                    ((f.category == 'Internal' or f.category == 'External') and 'script' in f.code_type.lower())]

        for f in findings:
            # For standalone files, the "Source" is the file itself. For references, it's the src attribute.
            src_url = f.snippet if f.category in ['Internal', 'External'] else "Self"
            is_remote = "Yes" if f.source_type == 'REMOTE' else "No"
            
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.code_type,
                src_url,
                f.start_line,
                is_remote
            ])
        self._create_sheet("External JS (Files)", headers, data)

    # --- CSS Sheets ---
    def _create_inline_css_sheet(self):
        """1. Inline CSS: Attributes (style=...)"""
        headers = ["File Path", "File Name", "Attribute", "Line", "Code Snippet"]
        data = []
        # Filter: CSS + Inline Source + NOT Style Block
        findings = [f for f in self.findings if f.category == 'CSS' and f.source_type == 'INLINE' and 'styleblock' not in f.code_type]
        
        for f in findings:
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.code_type,
                f.start_line,
                f.snippet
            ])
        self._create_sheet("Inline CSS (Attributes)", headers, data)

    def _create_internal_css_sheet(self):
        """2. Internal CSS: Style Blocks"""
        headers = ["File Path", "File Name", "Extracted File", "Line", "Code Snippet", "Full Code"]
        data = []
        # Filter: CSS + Inline Source + IS Style Block
        findings = [f for f in self.findings if f.category == 'CSS' and f.source_type == 'INLINE' and 'styleblock' in f.code_type]
        
        for f in findings:
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.bundled_file,
                f.start_line,
                f.snippet,
                f.full_code
            ])
        self._create_sheet("Internal CSS (Blocks)", headers, data)

    def _create_external_css_sheet(self):
        """3. External CSS: Local Files & Remote References"""
        headers = ["File Path", "File Name", "Reference Type", "Source/URL", "Line", "Is Remote?"]
        data = []
        
        # Filter: 
        # a) Category=CSS + Source=LOCAL (Standalone .css file)
        # b) Category=Internal/External + 'style'/'css' in type
        
        findings = [f for f in self.findings if 
                    (f.category == 'CSS' and f.source_type == 'LOCAL') or 
                    ((f.category == 'Internal' or f.category == 'External') and ('style' in f.code_type.lower() or 'css' in f.code_type.lower()))]

        for f in findings:
            src_url = f.snippet if f.category in ['Internal', 'External'] else "Self"
            is_remote = "Yes" if f.source_type == 'REMOTE' else "No"
            
            data.append([
                f.file_path,
                os.path.basename(f.file_path),
                f.code_type,
                src_url,
                f.start_line,
                is_remote
            ])
        self._create_sheet("External CSS (Files)", headers, data)

    def _create_sheet(self, title: str, headers: List[str], data_rows: List[List[str]]):
        ws = self.wb.create_sheet(title)
        
        # Header Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Write Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # Write Data
        for row_num, row_data in enumerate(data_rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                val_str = str(cell_value)
                # Excel cell limit is 32767 chars
                if len(val_str) > 32000:
                    val_str = val_str[:32000] + "..."
                cell.value = val_str
                
                # Align top for readability
                cell.alignment = Alignment(vertical="top", wrap_text=True if len(val_str) > 50 else False)
                cell.border = thin_border

        # Adjust column widths (basic heuristic)
        for col_num, col in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = col[0].column_letter
            
            # Don't auto-expand "Full Code" or "Snippet" too much
            header_val = ws.cell(row=1, column=col_num).value
            if header_val in ["Code Snippet", "Full Code", "Resource Path"]:
                ws.column_dimensions[column_letter].width = 50
                continue
                
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50: adjusted_width = 50
            if adjusted_width < 10: adjusted_width = 10
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_ajax_sheet(self):
        """Generates AJAX Code tab with detected AJAX calls and color coding."""
        # Columns: #, File Path, File Name, Category, Capability, Start Line, End Line, Endpoint/URL, Has Server Dependencies, Is Inline, Is Internal, Is External, Code Snippet, Full Code
        headers = ["#", "File Path", "File Name", "Category", "Capability", "Start Line", "End Line", 
                   "Endpoint/URL", "Has Server Dependencies", "Is Inline?", "Is Internal?", "Is External?", "Code Snippet", "Full Code"]
        data = []
        
        # Filter only AJAX-detected findings
        ajax_findings = [f for f in self.findings if f.ajax_detected]
        
        row_num = 1
        for f in ajax_findings:
            
            is_inline = "Yes" if f.source_type == 'INLINE' else "No"
            is_internal = "Yes" if f.source_type == 'LOCAL' else "No"
            is_external = "Yes" if f.source_type == 'REMOTE' else "No"
            
            # Use details list if available (Gold Standard)
            if getattr(f, 'ajax_details', None):
                for detail in f.ajax_details:
                    data.append([
                        row_num,
                        f.file_path,
                        os.path.basename(f.file_path),
                        detail.get('Category', 'Unknown'),   # New Col: Category (e.g. jQuery)
                        detail.get('Capability', 'Unknown'), # New Col: Capability (e.g. Real Call)
                        detail.get('Line', f.start_line),
                        f.end_line, # End line is approximate for the whole block
                        detail.get('Endpoint', 'Unknown'),
                        "Yes" if f.has_server_deps else "No",
                        is_inline,
                        is_internal,
                        is_external,
                        detail.get('Code_Snippet', f.snippet[:100]),
                        f.full_code # Full context is useful
                    ])
                    row_num += 1
            else:
                # Fallback for legacy Findings (safe guard)
                data.append([
                    row_num,
                    f.file_path,
                    os.path.basename(f.file_path),
                    f.ajax_pattern or "Unknown", # Category Fallback
                    "Unknown",                   # Capability Fallback
                    f.start_line,
                    f.end_line,
                    f.endpoint_url or "Unknown/Dynamic",
                    "Yes" if f.has_server_deps else "No",
                    is_inline,
                    is_internal,
                    is_external,
                    f.snippet,
                    f.full_code
                ])
                row_num += 1
        
        self._create_sheet("AJAX Code", headers, data)
        
        # Apply color coding to the AJAX tab
        if "AJAX Code" in self.wb.sheetnames:
            ws = self.wb["AJAX Code"]
            
            # Color code rows (starting from row 2, after header)
            for row in range(2, ws.max_row + 1):
                # Server Dependencies column (I = column 9)
                server_cell = ws.cell(row=row, column=9)
                if server_cell.value == "Yes":
                    server_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
                    server_cell.font = Font(color="9C0006")
                else:
                    server_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green
                    server_cell.font = Font(color="006100")

    def _create_refactoring_sheet(self, category_filter: str = None):
        """Generates Tab 4: Refactoring & Extraction Tracker (Developer Checklist).
           category_filter: 'JS' or 'CSS'
        """
        sheet_title = f"{category_filter} Refactoring" if category_filter else "Refactoring Tracker"
        
        headers = [
            "Ref ID", "Location", "Code Type", "Functionality", "Complexity", 
            "Extraction Status", "Target Filename", "Recommended Method", "Dev Notes"
        ]
        data = []
        
        row_num = 1
        for f in self.findings:
            # Filter Logic
            if category_filter:
                if f.category != category_filter:
                    continue
            elif f.category not in ['JS', 'CSS']:
                 continue
                 
            # Traffic Light Logic based on Severity and Complexity
            status = "🟢 Ready"
            method = "Move to separate file"
            
            if f.server_severity == "High":
                status = "🔴 Skipped (Blocked)"
                method = "Remove Server Dependencies First"
            elif f.server_severity == "Medium":
                status = "🟡 Skipped (Rewrite)"
                method = "Refactor Server Config (API)"
            elif f.complexity == "High":
                status = "🟡 Manual Review"
                method = "Convert to Component (Complex Logic)"
            elif f.functionality == "Page Glue":
                status = "🟢 Leave Inline"
                method = "None (Low Value)"
                
            # Generate Target Filename Recommendation (Strict Convention)
            # {OriginalFilePath}_{BlockType}_L{StartLine}-L{EndLine}.{Extension}
            clean_path = self._get_relative_path(f.file_path).replace(os.sep, "_").replace(".", "_")
            ext = ".js" if f.category == 'JS' else ".css"
            clean_type = "".join([c for c in f.code_type if c.isalnum()])
            target_name = f"{clean_path}_{clean_type}_L{f.start_line}-L{f.end_line}{ext}"
            
            # Ref ID
            ref_id = f"{f.category}-{row_num:04d}"
            
            
            data.append([
                ref_id,
                f"{os.path.basename(f.file_path)} : L{f.start_line}",
                f.code_type,
                f.functionality,
                f.complexity,
                status,
                target_name,
                method,
                "" # Dev Notes empty
            ])
            row_num += 1
            
        self._create_sheet(sheet_title, headers, data)
        
        # Color coding for Status
        if sheet_title in self.wb.sheetnames:
            ws = self.wb[sheet_title]
            for row in range(2, ws.max_row + 1):
                status_cell = ws.cell(row=row, column=6)
                val = status_cell.value
                if "Blocked" in val:
                    status_cell.fill = PatternFill("solid", fgColor="FFC7CE") # Red
                    status_cell.font = Font(color="9C0006")
                elif "Skipped" in val or "Manual" in val:
                    status_cell.fill = PatternFill("solid", fgColor="FFEB9C") # Yellow
                    status_cell.font = Font(color="9C6500")
                elif "Ready" in val:
                    status_cell.fill = PatternFill("solid", fgColor="C6EFCE") # Green
                    status_cell.font = Font(color="006100")

    def _create_legend_sheet(self):
        """Creates a Legend tab explaining the metrics."""
        ws = self.wb.create_sheet("Legend")
        
        headers = ["Term/Metric", "Definition", "Implication"]
        data = [
            ("Total Issues", "Sum of all Inline JS + Inline CSS + External Resource tags found.", "Indicates the total volume of work. High numbers = Heavy refactoring load."),
            ("Dynamic Code", "Presence of runtime code generation (eval, innerHTML, document.write).", "SECURITY RISK. Hard to measure complexity statically. Requires manual audit."),
            ("AJAX Calls", "Detected Asynchronous JavaScript patterns (local or remote).", "Indicates data flow. High count = Heavy API dependency."),
            ("Logic Density", "Score based on loops, conditionals, and logic structure.", "Low (<2) = Glue Code (keep inline?), High (>5) = Business Logic (Must Extract)."),
            ("Server Severity", "Presence of @Model, @ViewBag (Razor) or <% (ASP).", "High = Cannot move to .js file without rewriting logic to API/JSON."),
        ]
        
        # Header Style
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="000000")
            
        # Data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        # Widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60
