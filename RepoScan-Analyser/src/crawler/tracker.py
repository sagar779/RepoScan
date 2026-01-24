"""
Correlation Tracker - Generates Excel report with CSP analysis
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from urllib.parse import urlparse

class CorrelationTracker:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
    def generate_report(self, matches, new_findings, missing_findings, external_assets, output_path):
        self._create_summary(len(matches), len(new_findings), len(missing_findings), len(external_assets))
        self._create_csp_allowlist(matches, new_findings)
        self._create_external_sheet(external_assets)
        self._create_correlation_sheet(matches, new_findings, missing_findings)
        
        try:
            self.wb.save(output_path)
            print(f"Correlation report saved to: {output_path}")
        except Exception as e:
            print(f"Error saving report: {e}")

    def _get_csp_domain(self, endpoint):
        """Extract domain for CSP from endpoint URL"""
        if not endpoint or endpoint in ["Dynamic/Variable", "Server-Generated", "Unknown/Dynamic"]:
            return "Manual Review Required"
            
        if endpoint.startswith('/'):
            return "Self"
            
        try:
            parsed = urlparse(endpoint)
            if parsed.netloc:
                return parsed.netloc
            # Handle cases like "api.google.com/v1" without scheme
            if '.' in endpoint and '/' in endpoint:
                return endpoint.split('/')[0]
        except:
            pass
            
        return "Self (Assumed)"

    def _create_csp_allowlist(self, matches, new_findings):
        """Generates a unique list of domains for CSP configuration"""
        ws = self.wb.create_sheet("CSP Allowlist")
        
        # Aggregate domains
        domains = {} # domain -> {'count': 0, 'status': set()}
        
        all_items = matches + new_findings
        for item in all_items:
            endpoint = item.get('endpoint_url', '')
            domain = self._get_csp_domain(endpoint)
            
            if domain in ["Manual Review Required", "Self", "Self (Assumed)"]:
                continue
                
            if domain not in domains:
                domains[domain] = {'count': 0, 'status': set()}
            
            domains[domain]['count'] += 1
            # track if VERIFIED or NEW
            status = item.get('status', 'Unknown')
            domains[domain]['status'].add("Verified" if status == "VERIFIED" else "New")

        # Headers
        headers = ["Domain / Origin", "Frequency", "Source Status", "Suggested Directive"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="7030A0") # Purple
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            
        row = 2
        for domain, info in sorted(domains.items()):
            status_str = ", ".join(sorted(info['status']))
            
            ws.cell(row=row, column=1, value=domain)
            ws.cell(row=row, column=2, value=info['count'])
            ws.cell(row=row, column=3, value=status_str)
            ws.cell(row=row, column=4, value="connect-src")
            row += 1
            
        # Adjust widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_external_sheet(self, external_assets):
        """Creates a tab listing all external resources found (JS/CSS)"""
        ws = self.wb.create_sheet("External URLs")
        headers = ["Source Page", "Found External URL", "Type"]
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="000000") # Black
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            
        row = 2
        for asset in external_assets:
            ws.cell(row=row, column=1, value=asset['source_page'])
            ws.cell(row=row, column=2, value=asset['url'])
            ws.cell(row=row, column=3, value=asset['type'])
            row += 1
            
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 20

    def _create_summary(self, match_count, new_count, missing_count, external_count):
        ws = self.wb.create_sheet("Summary")
        ws['A1'] = "AJAX Correlation & CSP Report"
        ws['A1'].font = Font(bold=True, size=14)
        
        data = [
            ("Verified (Found in both)", match_count),
            ("New / Web Only (Dynamic)", new_count),
            ("Missing in Crawl (Unvisited/Dead)", missing_count),
            ("Total AJAX Patterns Tracked", match_count + new_count + missing_count),
            ("External Resources Found", external_count)
        ]
        
        for i, (label, count) in enumerate(data, 3):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=count)
            
        ws.column_dimensions['A'].width = 40

    def _create_correlation_sheet(self, matches, new_findings, missing):
        ws = self.wb.create_sheet("Correlation Matrix")
        headers = ["Status", "AJAX Type", "Extracted Endpoint", "CSP Domain", "Web URL", "Static File Path", "Code Snippet"]
        
        # Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            
        row = 2
        
        # 1. Verified Matches (Green)
        green_fill = PatternFill("solid", fgColor="C6EFCE")
        for m in matches:
            locations = ", ".join([loc['file'] for loc in m['static_locations']])
            endpoint = m.get('endpoint_url', 'Unknown')
            csp = self._get_csp_domain(endpoint)
            
            ws.cell(row=row, column=1, value="VERIFIED").fill = green_fill
            ws.cell(row=row, column=2, value=m['ajax_type'])
            ws.cell(row=row, column=3, value=endpoint)
            ws.cell(row=row, column=4, value=csp)
            ws.cell(row=row, column=5, value=m['dynamic_url'])
            ws.cell(row=row, column=6, value=locations)
            ws.cell(row=row, column=7, value=m['snippet'][:500])
            row += 1
            
        # 2. New Findings (Yellow - Warning)
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        for n in new_findings:
            endpoint = n.get('endpoint_url', 'Unknown')
            csp = self._get_csp_domain(endpoint)
            
            ws.cell(row=row, column=1, value="NEW_WEB_ONLY").fill = yellow_fill
            ws.cell(row=row, column=2, value=n['ajax_type'])
            ws.cell(row=row, column=3, value=endpoint)
            ws.cell(row=row, column=4, value=csp)
            ws.cell(row=row, column=5, value=n['dynamic_url'])
            ws.cell(row=row, column=6, value="N/A")
            ws.cell(row=row, column=7, value=n['snippet'][:500])
            row += 1
            
        # 3. Missing (Red - Alert)
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        for m in missing:
            locations = ", ".join([loc['file'] for loc in m['static_locations']])
            
            ws.cell(row=row, column=1, value="MISSING_IN_CRAWL").fill = red_fill
            ws.cell(row=row, column=2, value="Unknown")
            ws.cell(row=row, column=3, value="N/A")
            ws.cell(row=row, column=4, value="N/A") # CSP Domain unknown
            ws.cell(row=row, column=5, value="N/A")
            ws.cell(row=row, column=6, value=locations)
            ws.cell(row=row, column=7, value=m['snippet'])
            row += 1
            
        # Widths
        ws.column_dimensions['C'].width = 40 # Endpoint
        ws.column_dimensions['D'].width = 30 # CSP
        ws.column_dimensions['E'].width = 50 # Web URL
        ws.column_dimensions['F'].width = 50 # Static Path
        ws.column_dimensions['G'].width = 60 # Snippet
