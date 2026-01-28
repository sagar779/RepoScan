"""
Comparator - Correlates Static vs Dynamic findings
"""
import openpyxl
import re

def normalize_snippet(code):
    """Normalize code for comparison (ignore whitespace differences)"""
    return re.sub(r'\s+', ' ', code).strip()

class Comparer:
    def __init__(self, static_report_path):
        self.static_report_path = static_report_path
        self.static_snippets = {} # Hash -> List of Metadata
        self._load_static_report()
        
    def _load_static_report(self):
        try:
            wb = openpyxl.load_workbook(self.static_report_path, data_only=True)
            
            # Iterate through all sheets that might contain code
            for sheet_name in wb.sheetnames:
                if sheet_name in ["Summary", "Legend", "Output Manifest", "AJAX Code"]: continue
                
                ws = wb[sheet_name]
                # Identify columns dynamically
                headers = [cell.value for cell in ws[1]]
                
                try:
                    # Find column indices (1-based for openpyxl)
                    if 'Code Snippet' not in headers or 'File Path' not in headers:
                        continue
                        
                    snippet_col_idx = headers.index('Code Snippet') + 1
                    file_col_idx = headers.index('File Path') + 1
                    print(f"Loading static findings from '{ws.title}'...")
                except ValueError:
                    continue
    
                # Read rows
                for row in range(2, ws.max_row + 1):
                    snippet = ws.cell(row=row, column=snippet_col_idx).value
                    filepath = ws.cell(row=row, column=file_col_idx).value
                    
                    if snippet:
                        norm = normalize_snippet(str(snippet))
                        if norm not in self.static_snippets:
                            self.static_snippets[norm] = []
                        self.static_snippets[norm].append({'file': filepath, 'row': row})
            
            print(f"Loaded {len(self.static_snippets)} unique static snippets.")

        except Exception as e:
            print(f"Error loading static report: {e}")

    def correlate(self, dynamic_findings):
        """
        Compare dynamic CodeSnippet objects against loaded static snippets.
        Returns: Tuple (matches, new_findings, missing_static_keys)
        """
        matches = []
        new_findings = []
        matched_keys = set()
        
        for finding in dynamic_findings:
            norm = normalize_snippet(finding.snippet)
            
            if norm in self.static_snippets:
                # Match Found!
                matched_static = self.static_snippets[norm] # list of {file, row}
                match_record = {
                    'status': 'VERIFIED',
                    'dynamic_url': finding.file_path, # URL is stored in file_path for dynamic
                    'static_locations': matched_static,
                    'snippet': finding.snippet,
                    'ajax_type': finding.ajax_pattern,
                    'endpoint_url': finding.endpoint_url
                }
                matches.append(match_record)
                matched_keys.add(norm)
            else:
                # Found on Web but NOT in Static
                new_record = {
                    'status': 'NEW_WEB_ONLY',
                    'dynamic_url': finding.file_path,
                    'snippet': finding.snippet,
                    'ajax_type': finding.ajax_pattern,
                    'endpoint_url': finding.endpoint_url
                }
                new_findings.append(new_record)
                
        # Calculate missing (Static items that were never seen on web)
        missing_findings = []
        for key, locations in self.static_snippets.items():
            if key not in matched_keys:
                missing_findings.append({
                    'status': 'MISSING_IN_CRAWL',
                    'static_locations': locations,
                    'snippet': key[:100] + '...' # Truncate for report
                })
                
        return matches, new_findings, missing_findings
